from collections import defaultdict
import os
import psycopg2
from openpyxl import Workbook
import boto3
import requests


def is_cast_to_int(value):
    try:
        int(value)
        return True
    except ValueError:
        return False


start_timestamp = '2024-01-01T00:00:00.000Z'
end_timestamp = '2024-06-21T00:00:00.000Z'


def fetchReportRunAuditLogEvents(start_timestamp, end_timestamp):
    dynamodb = boto3.client('dynamodb')
    table_name = 'bichard-7-production-audit-log-events'
    index_name = 'eventCodeIndex'
    key_condition_expression = 'eventCode = :code and #ts BETWEEN :start_time AND :end_time'
    expression_attribute_names = {'#ts': 'timestamp'}
    expression_attribute_values = {
        ':code': {'S': 'report-run'},
        ':start_time': {'S': start_timestamp},
        ':end_time': {'S': end_timestamp}
    }

    items = []
    last_evaluated_key = None

    while True:
        if last_evaluated_key:
            response = dynamodb.query(
                TableName=table_name,
                IndexName=index_name,
                KeyConditionExpression=key_condition_expression,
                ExpressionAttributeNames=expression_attribute_names,
                ExpressionAttributeValues=expression_attribute_values,
                ExclusiveStartKey=last_evaluated_key
            )
        else:
            response = dynamodb.query(
                TableName=table_name,
                IndexName=index_name,
                KeyConditionExpression=key_condition_expression,
                ExpressionAttributeNames=expression_attribute_names,
                ExpressionAttributeValues=expression_attribute_values
            )

        items.extend(response.get('Items', []))
        last_evaluated_key = response.get('LastEvaluatedKey')

        if not last_evaluated_key:
            break

    return items


# Fetch forces data
url = 'https://raw.githubusercontent.com/ministryofjustice/bichard7-next-data/main/output-data/data/forces.json'
response = requests.get(url)
forces = response.json()

forcesDict = {}
for f in forces:
    forcesDict[f['code']] = f['name']


def orderCodes(codes):
    code_list = [code for code in codes.split(',')]
    sorted_codes = sorted(code_list)
    return ','.join([code for code in sorted_codes])


def fetchUsers(conn):
    users = {}
    cur = conn.cursor()
    cur.execute(f"""SELECT username, visible_courts, visible_forces, email
                FROM "br7own"."users"
                LIMIT 10000 OFFSET 0
                """)
    results = cur.fetchall()
    for user in results:
        users[user[0]] = {
            "courts": user[1],
            "forces": user[2],
            "email": user[3]
        }
    return users


verbose_events = fetchReportRunAuditLogEvents(start_timestamp, end_timestamp)
events = []

conn_string = os.getenv('CONNECTION_STRING')
conn = psycopg2.connect(conn_string)
users = fetchUsers(conn)

for event in verbose_events:
    if not event['user']['S'] in users:
        print(f"Cannot find user {event['user']}")
    else:
        list_of_forces = orderCodes(
            users[event['user']['S']]['forces']).split(',')[0]
        if (is_cast_to_int(list_of_forces)):
            forceCode = str(int(list_of_forces)).zfill(2)
            force = forcesDict[forceCode]
        else:
            force: list_of_forces

        events.append({
            "report": event['attributes']['M']['Report ID']['S'],
            "user": event['user']['S'],
            "output": event['attributes']['M']['Output Format']['S'],
            "timeStamp": event['timestamp']['S'],
            "force": force,
            "email": users[event['user']['S']]['email']
        })

forces_counts = defaultdict(int)
court_counts = defaultdict(int)
report_counts = defaultdict(int)
force_report_counts = defaultdict(lambda: defaultdict(int))
user_reports_counts = defaultdict(lambda: defaultdict(int))

for event in events:
    report = event['report']
    force = event['force']
    user = event['user']

    # Count report types
    report_counts[report] += 1

    # Count user-specific report types
    force_report_counts[report][force] += 1
    user_reports_counts[report][user] += 1


wb = Workbook()
sheet = wb.active

sheet.cell(row=1, column=1, value="Report")
sheet.cell(row=1, column=2, value="Count")

row_num = 2
for report, count in report_counts.items():
    sheet.cell(row=row_num, column=1, value=report)
    sheet.cell(row=row_num, column=2, value=count)
    row_num += 1

sheet.cell(row=1, column=4, value="Report")
sheet.cell(row=1, column=5, value="Force")
sheet.cell(row=1, column=6, value="Count")

row_num = 2
for report, report_data in force_report_counts.items():
    for force, count in report_data.items():
        sheet.cell(row=row_num, column=4, value=report)
        sheet.cell(row=row_num, column=5, value=force)
        sheet.cell(row=row_num, column=6, value=count)
        row_num += 1

sheet.cell(row=1, column=8, value="Report")
sheet.cell(row=1, column=9, value="Force")
sheet.cell(row=1, column=10, value="User")
sheet.cell(row=1, column=11, value="Count")

row_num = 2
for report, report_data in user_reports_counts.items():
    for user, count in report_data.items():
        list_of_forces = orderCodes(users[user]['forces']).split(',')[0]
        if (is_cast_to_int(list_of_forces)):
            forceCode = str(int(list_of_forces)).zfill(2)
            force = forcesDict[forceCode]
        else:
            force = list_of_forces
        sheet.cell(row=row_num, column=8, value=report)
        sheet.cell(row=row_num, column=9, value=force)
        sheet.cell(row=row_num, column=10, value=user)
        sheet.cell(row=row_num, column=11, value=count)
        row_num += 1

wb.save("report_analysis_results.xlsx")
